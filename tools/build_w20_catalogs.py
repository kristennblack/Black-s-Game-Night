from __future__ import annotations
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import json, re, hashlib, math, html

ROOT=Path('/mnt/data/bfgn_w20')
WB=Path('/mnt/data/Black_Family_Game_Night_V1_Master_Catalog_Plan.xlsx')
LEGACY_HOME=Path('/mnt/data/w19_home_legacy.json')
LEGACY_WEAR=Path('/mnt/data/w19_wear_legacy.json')
wb=load_workbook(WB,data_only=True,read_only=True)

def rows(sheet, start=5):
    ws=wb[sheet]
    hdr=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]
    out=[]
    for r in range(start,ws.max_row+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if not vals[0]: continue
        out.append(dict(zip(hdr,vals)))
    return out

home_alloc={r['Category']:int(r['Target Count']) for r in rows('Home Allocation') if isinstance(r.get('Target Count'),(int,float))}
wear_alloc={r['Category']:int(r['Target Count']) for r in rows('Wearable Allocation') if isinstance(r.get('Target Count'),(int,float))}
home_cols=rows('Home Collections')
wear_cols=rows('Wearable Collections')
home_flag=rows('Flagship Home 100')[:100]
wear_flag=rows('Flagship Wearables 100')[:100]
legacy_home=json.loads(LEGACY_HOME.read_text())
legacy_wear=json.loads(LEGACY_WEAR.read_text())

home_cat_map={
'Beds':'Beds & Bedroom Furniture','Seating':'Seating','Tables':'Tables & Desks','Storage':'Storage','Lighting':'Lighting',
'Electronics':'Electronics & Entertainment','Plants':'Clutter & Detail Props','Decorations':'Clutter & Detail Props','Collectibles':'Clutter & Detail Props',
'Toys & Hobbies':'Clutter & Detail Props','Games':'Specialty / Family / Interactive / Hero','Pet Items':'Clutter & Detail Props','Rugs':'Rugs & Soft Decor',
'Wall Decor':'Wall Decor & Pictures','Architecture':'Architectural Finishes','Wallpaper':'Architectural Finishes','Flooring':'Architectural Finishes',
'Ceiling & Trim':'Architectural Finishes','Windows & Doors':'Windows & Doors','Special Effects':'Specialty / Family / Interactive / Hero'
}
wear_cat_by_slot={
'hat':'Hats / Headwear','glasses':'Glasses / Face Accessories','headset':'Hats / Headwear','jewelry':'Jewelry','top':'Tops',
'badge':'Scarves / Belts / Wrist Accessories','accessory':'Scarves / Belts / Wrist Accessories','hair':'Wigs / Hairstyles'
}

def slug(s):
    s=str(s).lower().replace('’',"'")
    s=re.sub(r"[^a-z0-9]+",'-',s).strip('-')
    return s[:120]

def seed(s):
    return int(hashlib.sha256(str(s).encode()).hexdigest()[:12],16)

def choose(seq,n): return seq[n%len(seq)]

def collection_meta(cols): return {r['Collection']:r for r in cols}
home_col_meta=collection_meta(home_cols); wear_col_meta=collection_meta(wear_cols)
home_col_names=list(home_col_meta); wear_col_names=list(wear_col_meta)

# Category-aware collection pools keep generated content coherent.
home_pools={
'Clutter & Detail Props':['Rustic Pine','Old Lodge','Cozy Cabin','Heritage Farmhouse','Workshop Life','Country Kitchen','Game Night','Kids & Play','Pet Corner','Outdoorsman','Vintage Camp','Handmade Homestead','John’s Workshop','Holly Pop','Logan Outdoors','Papa’s Old Shop','Nana’s Classics','James & Dorothy Home'],
'Wall Decor & Pictures':['Old Lodge','Mountain Retreat','Lakeside Weekend','Western Cabin','Outdoorsman','Collector’s Den','John’s Workshop','Lizzy Ballet','Holly Pop','Nana’s Classics','James & Dorothy Home'],
'Architectural Finishes':['Rustic Pine','Old Lodge','Cozy Cabin','Heritage Farmhouse','Mountain Retreat','Lakeside Weekend','Modern Lodge','Lodge Luxury','Western Cabin','Vintage Camp','Handmade Homestead','Kristen’s Cabin Chic','Vanessa Modern'],
'Windows & Doors':['Rustic Pine','Old Lodge','Heritage Farmhouse','Mountain Retreat','Lakeside Weekend','Modern Lodge','Lodge Luxury','Western Cabin','Vintage Camp','Handmade Homestead'],
'Lighting':['Rustic Pine','Old Lodge','Cozy Cabin','Heritage Farmhouse','Workshop Life','Fireside Living','Mountain Retreat','Modern Lodge','Lodge Luxury','Western Cabin','Vintage Camp','Kristen’s Cabin Chic','Vanessa Modern','Nana’s Classics'],
'Beds & Bedroom Furniture':['Rustic Pine','Old Lodge','Cozy Cabin','Heritage Farmhouse','Mountain Retreat','Lakeside Weekend','Modern Lodge','Lodge Luxury','Western Cabin','Vintage Camp','Handmade Homestead','Kristen’s Cabin Chic','Holly Pop','Lizzy Ballet','Logan Outdoors','Vanessa Modern','Nana’s Classics'],
'Seating':['Rustic Pine','Old Lodge','Cozy Cabin','Heritage Farmhouse','Fireside Living','Mountain Retreat','Lakeside Weekend','Modern Lodge','Lodge Luxury','Western Cabin','Vintage Camp','Handmade Homestead','Vanessa Modern','Papa’s Old Shop','Nana’s Classics','James & Dorothy Home'],
'Tables & Desks':['Rustic Pine','Old Lodge','Heritage Farmhouse','Workshop Life','Country Kitchen','Modern Lodge','Western Cabin','Handmade Homestead','John’s Workshop','Kristen’s Cabin Chic','Vanessa Modern','Papa’s Old Shop'],
'Storage':['Rustic Pine','Old Lodge','Heritage Farmhouse','Workshop Life','Country Kitchen','Modern Lodge','Game Night','Kids & Play','Pet Corner','Collector’s Den','Outdoorsman','Handmade Homestead','John’s Workshop','Holly Pop','Logan Outdoors'],
'Rugs & Soft Decor':['Cozy Cabin','Heritage Farmhouse','Fireside Living','Mountain Retreat','Lakeside Weekend','Lodge Luxury','Kids & Play','Pet Corner','Western Cabin','Vintage Camp','Handmade Homestead','Kristen’s Cabin Chic','Holly Pop','Lizzy Ballet','Nana’s Classics'],
'Kitchen & Bath Utility Decor':['Heritage Farmhouse','Country Kitchen','Modern Lodge','Handmade Homestead','Kristen’s Cabin Chic','Vanessa Modern','Nana’s Classics','James & Dorothy Home'],
'Electronics & Entertainment':['Modern Lodge','Game Night','Movie Cabin','Kids & Play','Collector’s Den','Vanessa Modern'],
'Outdoor / Porch / Deck':['Rustic Pine','Lakeside Weekend','Outdoorsman','Vintage Camp','Handmade Homestead','Logan Outdoors','James & Dorothy Home'],
'Specialty / Family / Interactive / Hero':['Lodge Luxury','Game Night','Movie Cabin','Collector’s Den','John’s Workshop','Kristen’s Cabin Chic','Holly Pop','Lizzy Ballet','Logan Outdoors','Vanessa Modern','Papa’s Old Shop','Nana’s Classics','James & Dorothy Home','Cabin Celebration']
}

wear_pools={cat:[r['Collection'] for r in wear_cols if r['Collection'] not in ('Family Celebration',)] for cat in wear_alloc}
wear_pools.update({
'Tops':['Cabin Casual','Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Cozy Pajamas','Glam Night','Formal Family','Retro Closet','Camp Life','Sporty','Game Night','Cute & Playful','Winter Lodge','Summer Weekend','John Workwear','Kristen Cabin Chic','Holly Cute Pop','Lizzy Ballet Glam','Logan Trail & Fishing','Vanessa Modern','Papa Rugged','Nana Classic'],
'Bottoms':['Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Cozy Pajamas','Formal Family','Camp Life','Sporty','Summer Weekend','Vanessa Modern'],
'Dresses / Skirts / One-Piece':['Western Night','Cozy Pajamas','Glam Night','Formal Family','Retro Closet','Cute & Playful','Festival','Summer Weekend','Kristen Cabin Chic','Holly Cute Pop','Lizzy Ballet Glam','Vanessa Modern','Nana Classic'],
'Jackets / Coats / Outerwear':['Cabin Casual','Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Glam Night','Formal Family','Retro Closet','Camp Life','Sporty','Festival','Winter Lodge','John Workwear','Kristen Cabin Chic','Logan Trail & Fishing','Vanessa Modern','Papa Rugged'],
'Footwear':['Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Cozy Pajamas','Glam Night','Formal Family','Retro Closet','Camp Life','Sporty','Cute & Playful','Winter Lodge','Summer Weekend','Lizzy Ballet Glam'],
'Hats / Headwear':['Cabin Casual','Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Cozy Pajamas','Glam Night','Formal Family','Retro Closet','Camp Life','Sporty','Cute & Playful','Festival','Costume Trunk','Winter Lodge','Summer Weekend','John Workwear','Holly Cute Pop','Lizzy Ballet Glam','Logan Trail & Fishing','Papa Rugged','Pet Party'],
'Wigs / Hairstyles':['Glam Night','Formal Family','Retro Closet','Cute & Playful','Festival','Costume Trunk','Hair & Wig Bar','Lizzy Ballet Glam','Vanessa Modern','Nana Classic'],
'Glasses / Face Accessories':['Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Glam Night','Formal Family','Retro Closet','Sporty','Cute & Playful','Festival','Costume Trunk','Silly Filters','John Workwear','Vanessa Modern','Papa Rugged','Nana Classic'],
'Snapchat-Style Filters / Effects':['Cute & Playful','Festival','Costume Trunk','Silly Filters','Fantasy Filters','Pet Party','Family Celebration'],
'Jewelry':['Western Night','Glam Night','Formal Family','Retro Closet','Festival','Jewelry Box','Kristen Cabin Chic','Holly Cute Pop','Lizzy Ballet Glam','Vanessa Modern','Nana Classic','Pet Party'],
'Bags / Back Items':['Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Glam Night','Formal Family','Retro Closet','Camp Life','Sporty','Cute & Playful','Festival','Costume Trunk','Summer Weekend','Kristen Cabin Chic','Holly Cute Pop','Logan Trail & Fishing','Vanessa Modern'],
'Scarves / Belts / Wrist Accessories':['Cabin Casual','Country Weekend','Western Night','Work Crew','Outdoors & Fishing','Glam Night','Formal Family','Retro Closet','Camp Life','Sporty','Cute & Playful','Festival','Jewelry Box','Winter Lodge','John Workwear','Kristen Cabin Chic','Logan Trail & Fishing','Papa Rugged','Nana Classic','Pet Party'],
'Costumes / Novelty Clothing':['Western Night','Work Crew','Outdoors & Fishing','Cozy Pajamas','Game Night','Cute & Playful','Festival','Costume Trunk','Pet Party','Family Celebration'],
'Wings / Tails / Ears / Horns / Attachments':['Cute & Playful','Festival','Costume Trunk','Silly Filters','Fantasy Filters','Pet Party','Family Celebration']
})

home_nouns={
'Clutter & Detail Props':['Book Stack','Mug Set','Basket','Candle Cluster','Tool Tray','Snack Caddy','Boot Pair','Jar Set','Blanket Basket','Remote Tray','Game Box','Tackle Tin','Dog Toy Bin','Pottery Set','Photo Frame Set','Key Bowl','Lantern Pair','Magazine Rack','Craft Caddy','Firewood Crate','Tea Tray','Record Stack','Desk Organizer','Fishing Creel','Board Game Stack'],
'Wall Decor & Pictures':['Framed Landscape','Round Mirror','Wall Clock','Timber Sign','Shadowbox','Tapestry','Photo Collage','Carved Panel','Map Print','Pressed Flower Frame','Metal Wall Art','Antler Plaque','Ballet Print','Workshop Diagram','Lake Chart','Family Frame','Quilted Wall Hanging','Cabin Motto Sign'],
'Architectural Finishes':['Wallpaper','Wall Paneling','Log Wall','Stone Accent Wall','Plank Ceiling','Beam Set','Flooring','Brick Accent','Wainscot','Trim Package','Mural Wall','Plaid Wallcloth','Limewash Wall','Board-and-Batten Wall','Tin Ceiling'],
'Windows & Doors':['Picture Window','Arched Window','Casement Window','Bay Window','Round Window','Dormer Window','French Door','Barn Door','Panel Door','Arched Door','Pocket Door','Double Cabin Door','Dutch Door','Glass Door','Hidden Door'],
'Lighting':['Table Lamp','Floor Lamp','Chandelier','Pendant Light','Wall Sconce','Lantern','String Lights','Desk Lamp','Reading Lamp','Ceiling Fixture','Task Light','Mason Jar Light','Antler Light','Branch Lamp','Picture Light'],
'Beds & Bedroom Furniture':['Platform Bed','Canopy Bed','Spindle Bed','Sleigh Bed','Daybed','Bunk Bed','Iron Bed','Storage Bed','Upholstered Bed','Log Bed','Trundle Bed','Four-Poster Bed','Cabin Cot','Headboard Bed','Floating Bed'],
'Seating':['Club Chair','Reading Chair','Rocking Chair','Sling Chair','Bean Bag','Bench','Loveseat','Sofa','Window Seat','Recliner','Barrel Chair','Dining Chair','Desk Chair','Hanging Chair','Chaise'],
'Tables & Desks':['Dining Table','Farm Table','Live-Edge Table','Coffee Table','Side Table','Writing Desk','Workshop Desk','Secretary Desk','Corner Desk','Game Table','Vanity Desk','Console Table','Breakfast Table','Craft Table'],
'Storage':['Dresser','Wardrobe','Trunk','Bookcase','Cabinet','Storage Bench','Cubby Shelf','Glass Cabinet','Ladder Shelf','Locker','Nightstand','Media Cabinet','Pantry Shelf','Display Hutch','Toy Chest'],
'Rugs & Soft Decor':['Area Rug','Runner Rug','Quilt','Throw Blanket','Curtain Set','Pillow Pair','Bedding Set','Window Valance','Pouf','Floor Cushion','Wool Throw','Braided Rug','Hide-Style Rug','Bench Cushion','Canopy Drapes'],
'Kitchen & Bath Utility Decor':['Dish Rack','Towel Ladder','Soap Tray','Utensil Crock','Pantry Canister Set','Bath Caddy','Toilet Paper Stand','Laundry Hamper','Spice Rack','Knife Block','Coffee Station','Tea Station','Sink Caddy','Towel Ring','Cookbook Stand'],
'Electronics & Entertainment':['Wall TV','Console TV','Projector','Record Player','Radio','Game Console Dock','Speaker Pair','Movie Projector','Arcade Cabinet','Digital Photo Frame','Media Hub','Portable Radio','Karaoke Speaker','Turntable','Mini Theater Screen'],
'Outdoor / Porch / Deck':['Porch Chair','Camp Chair','Deck Bench','Planter Box','Lantern Post','Boot Rack','Grill Cart','Picnic Table','Firewood Rack','Outdoor Rug','Cooler Stand','Deck Side Table','Hammock','Porch Swing','Garden Bench'],
'Specialty / Family / Interactive / Hero':['Stone Fireplace','Game Vault','Aquarium Divider','Hidden Bookcase','Trophy Display','Interactive Dog Station','Secret Passage Door','Music Corner','Photo Booth','Ballet Barre Wall','Workshop Command Wall','Collector Cabinet','Movie Snack Bar','Cabin Bar Cart','Signature Showcase']
}
wear_nouns={
'Tops':['Flannel Overshirt','Knit Sweater','Snap Shirt','Western Shirt','Fishing Shirt','Hoodie','Wrap Sweater','Cardigan','Camp Tee','Game Jersey','Henley','Thermal Top','Plaid Shirt','Ribbed Tank','Work Shirt'],
'Bottoms':['Utility Pant','Straight-Leg Denim','Cargo Pant','Trail Short','Jogger','Corduroy Pant','Wide-Leg Trouser','Western Jean','Lounge Pant','Canvas Short','Pleated Trouser','Denim Short'],
'Dresses / Skirts / One-Piece':['Sundress','Velvet Dress','Western Skirt','Wrap Dress','Shirt Dress','Sweater Dress','Ballet Skirt','Denim Dress','Jumpsuit','Overalls','Formal Gown','Tea Dress'],
'Jackets / Coats / Outerwear':['Work Jacket','Parka','Windbreaker','Shacket','Moto Jacket','Sequin Jacket','Denim Jacket','Wool Coat','Puffer Vest','Rain Shell','Western Duster','Fleece Jacket','Quilted Vest','Bomber Jacket'],
'Footwear':['Hiking Boot','Western Boot','Slippers','Trail Sneaker','Ballet Flat','Work Boot','Chelsea Boot','High-Top Sneaker','Loafer','Sandal','Rain Boot','Moccasin','Dress Shoe','Cozy Sock Slipper'],
'Hats / Headwear':['Shop Cap','Sun Hat','Bucket Hat','Western Hat','Party Crown','Earmuffs','Beanie','Toque','Headband','Beret','Newsboy Cap','Trucker Cap','Cowboy Hat','Tiara','Bandana Headwrap','Flower Crown'],
'Wigs / Hairstyles':['Long Waves Wig','Short Bob Wig','Braided Pigtails','Curly Updo','Spiky Wig','Ombre Wig','Blue Streak Wig','Long Straight Wig','Mohawk Wig','Silver Bob Wig','Big Curls Wig','Ballet Bun Wig','Shag Wig','Pixie Wig'],
'Glasses / Face Accessories':['Aviator Shades','Round Glasses','Heart Glasses','Mask','Star Glasses','Cat-Eye Glasses','Reading Glasses','Ski Goggles','Funny Moustache','Monocle','Sport Shades','Clear Frames','Bandit Mask','Retro Shades'],
'Snapchat-Style Filters / Effects':['Puppy Nose Filter','Deer Antlers Filter','Heart Crown Filter','Sparkle Glow Filter','Rainbow Tears Filter','Butterfly Aura Filter','Sunflower Glasses Filter','Clown Nose Filter','Pixel Glasses Filter','Love Eyes Filter','Angel Glow Filter','Snowflake Filter','Freckle Glow Filter','Comic Brows Filter','Firefly Face Filter'],
'Jewelry':['Silver Necklace','Turquoise Pendant','Gold Chain','Choker','Pearl Necklace','Hoop Earrings','Drop Earrings','Stud Earrings','Charm Bracelet','Leather Bracelet','Classic Watch','Statement Ring','Pendant Necklace','Beaded Bracelet'],
'Bags / Back Items':['Canvas Backpack','Structured Mini Bag','Trail Pack','Crossbody Bag','Fishing Sling','Glam Clutch','Tool Backpack','Duffel Bag','Mini Satchel','Quilted Backpack','Costume Quiver','Camera Bag'],
'Scarves / Belts / Wrist Accessories':['Plaid Scarf','Bandana','Leather Belt','Bracelet Set','Classic Watch','Western Belt','Knit Scarf','Silk Scarf','Wrist Cuff','Utility Belt','Ballet Ribbon Wrap','Camp Patch Belt','Charm Wristband'],
'Costumes / Novelty Clothing':['Firefighter Costume','Storybook Cape','Dinosaur Suit','Cow Costume','Pirate Outfit','Royal Outfit','Astronaut Suit','Camp Ranger Costume','Goat Costume','Card Shark Suit','Ballerina Costume','Mechanic Costume','Fairy Outfit','Retro Ski Suit'],
'Wings / Tails / Ears / Horns / Attachments':['Bunny Ears','Devil Horns','Halo','Fairy Wings','Dinosaur Hood','Cat Ears','Fox Tail','Angel Wings','Butterfly Wings','Dragon Tail','Goat Horns','Bear Ears','Star Wings','Rainbow Wings']
}

materials=['Pine','Oak','Walnut','Iron','Leather','Canvas','Wool','Brass','Stone','Birch','Cedar','Linen','Plaid','Enamel','Rattan','Copper','Maple','Steel','Velvet','Ceramic']
features=['Carved','Slatted','Turned','Handmade','Quilted','Weathered','Layered','Framed','Inset','Woven','Riveted','Scalloped','Paneled','Braided','Distressed','Polished','Tapered','Curved','Chunky','Slimline','Patchwork','Embroidered','Stitched','Sculptural']
colors=['Forest','Charcoal','Burgundy','Cream','Tan','Navy','Rose','Gold','Silver','Teal','Moss','Rust','Cognac','Black','Blush','Sky','Olive','Plum']

rarities=['Common','Uncommon','Rare','Premium','Legendary','Family Signature']
price_base={'Common':60,'Uncommon':110,'Rare':220,'Premium':450,'Legendary':900,'Family Signature':650,'Event Exclusive':0}

def rarity_for(index, collection, role=''):
    if 'Signature' in str(role) or home_col_meta.get(collection,{}).get('Tier')=='Signature' or wear_col_meta.get(collection,{}).get('Tier')=='Signature': return 'Family Signature'
    if 'Hero' in str(role):
        return 'Legendary' if index%3==0 else 'Premium'
    x=index%100
    if x<44:return 'Common'
    if x<70:return 'Uncommon'
    if x<86:return 'Rare'
    if x<95:return 'Premium'
    return 'Legendary'

def source_for(rarity,index,role=''):
    if rarity=='Family Signature' and index%4==0:return 'Achievement Reward'
    if 'Event' in str(role):return 'Event Reward'
    if rarity=='Legendary' and index%7==0:return 'Achievement Reward'
    if index%23==0:return 'Win in Arcade'
    return 'Buy with Game Night Tokens'

def token_price(rarity,index,source):
    if source!='Buy with Game Night Tokens': return 0
    base=price_base[rarity]
    bump=(index%7)*10
    return base+bump

def home_surface(cat,noun):
    if cat in ('Wall Decor & Pictures',): return 'Wall'
    if cat=='Architectural Finishes':
        if any(k in noun for k in ['Flooring']):return 'Floor'
        return 'Wall'
    if cat=='Windows & Doors':return 'Wall'
    if cat=='Lighting' and any(k in noun for k in ['Sconce','Picture','Ceiling','Chandelier','Pendant']):return 'Wall; Ceiling'
    return 'Floor'

def home_subcat(cat,noun): return noun

def footprint(cat,i):
    dims={
      'Clutter & Detail Props':(0.7,0.7),'Wall Decor & Pictures':(1.2,0.3),'Architectural Finishes':(2.0,0.2),'Windows & Doors':(2.0,0.4),
      'Lighting':(1.0,1.0),'Beds & Bedroom Furniture':(2.2,3.2),'Seating':(1.7,1.7),'Tables & Desks':(2.1,1.4),'Storage':(1.8,0.8),
      'Rugs & Soft Decor':(2.4,2.4),'Kitchen & Bath Utility Decor':(0.8,0.7),'Electronics & Entertainment':(1.7,0.7),'Outdoor / Porch / Deck':(2.0,1.5),
      'Specialty / Family / Interactive / Hero':(2.4,1.8)}
    w,d=dims[cat];return round(w+(i%3)*.2,1),round(d+((i//3)%3)*.2,1)

def new_home_record(item_id,name,cat,collection,rarity,role,concept,notes,index):
    source=source_for(rarity,index,role); price=token_price(rarity,index,source); w,d=footprint(cat,index)
    noun=home_subcat(cat, choose(home_nouns[cat],index))
    interactive=('Interactive' in str(role)) or (cat in ['Lighting','Windows & Doors','Electronics & Entertainment','Specialty / Family / Interactive / Hero'] and index%5==0)
    return {
      'Item ID':item_id,'Item Name':name,'Category':cat,'Subcategory':noun,'Collection':collection,'Rarity':rarity,
      'Source Type':source,'Source Game':'Arcade Rotation' if source=='Win in Arcade' else '',
      'Visible Unlock Condition':'Available in the Cabin Shop' if source=='Buy with Game Night Tokens' else ('Earn through play' if source!='Event Reward' else 'Family event reward'),
      'Internal Unlock Condition':'catalog-w20','Token Price':price,'Equivalent Token Value':price or price_base.get(rarity,300),'Salvage %':0.25,
      'Giftable':'No' if source=='Achievement Reward' else 'Yes','Unlimited Placement':'Yes','Account-Earned Only':'Yes' if source=='Achievement Reward' else 'No','Secret':'No',
      'Footprint W':w,'Footprint D':d,'Height Band':'Tall' if cat in ['Beds & Bedroom Furniture','Storage','Windows & Doors','Lighting'] else 'Medium',
      'Placement Surface':home_surface(cat,noun),'Room Types':'Bedroom; Common Space; Specialty Rooms','Rotation':'90° steps',
      'Style Tags':f'{collection}; {rarity}; W20; {role or "Standard"}','Animation / VFX':'Interactive animation' if interactive else 'None',
      'Audio / Ambience':'Contextual' if interactive and index%3==0 else 'None','Future Interaction':concept if interactive else notes,
      'Collection Completion Set':collection,'3D Production Notes':f'W20 identity seed {seed(item_id)}. Preserve silhouette/material identity across browse, preview and game mesh.',
      'Visual Concept':concept,'Production Notes':notes,'Hero':'Yes' if ('Hero' in str(role) or rarity=='Legendary') else 'No','Interactive':'Yes' if interactive else 'No',
      'Art Seed':seed(item_id),'Catalog Version':20
    }

# Build home catalog, preserving legacy IDs but normalizing into W20 taxonomy.
home=[]; used_ids=set(); used_names=set(); cat_counts=Counter()
for i,x in enumerate(legacy_home):
    y=dict(x); oldcat=y.get('Category'); cat=home_cat_map.get(oldcat,oldcat)
    y['Category']=cat; y['Subcategory']=y.get('Subcategory') or oldcat; y['Catalog Version']=20; y['Legacy Preserved']='Yes'; y['Art Seed']=seed(y['Item ID'])
    y['Hero']='Yes' if y.get('Rarity') in ('Family Legendary','Legendary') else 'No'; y['Interactive']='Yes' if y.get('Future Interaction') and y.get('Future Interaction')!='None' else 'No'
    used_ids.add(y['Item ID']); used_names.add(y['Item Name'].casefold()); cat_counts[cat]+=1; home.append(y)

for idx,f in enumerate(home_flag,1):
    cat=f['Category']; name=f['Name']; collection=f['Collection']; role=f['Role'] or ''
    if name.casefold() in used_names: continue
    iid=f"home-flagship-{str(f['Item ID']).lower()}-{slug(name)}"
    rar=f['Rarity'] or rarity_for(idx,collection,role)
    rec=new_home_record(iid,name,cat,collection,rar,role,f['Visual Concept'] or '',f['Production / Fit Notes'] or '',idx+400)
    home.append(rec);used_ids.add(iid);used_names.add(name.casefold());cat_counts[cat]+=1

for cat,target in home_alloc.items():
    pool=home_pools[cat]; nouns=home_nouns[cat]
    n=0
    while cat_counts[cat]<target:
        n+=1; global_i=len(home)+1; collection=choose(pool,global_i+n*3); noun=choose(nouns,n+global_i)
        material=choose(materials,global_i*3+n); feature=choose(features,global_i*5+n*2)
        # Collection prefix is omitted when it would make names unwieldy; material/feature carry identity.
        if cat in ('Architectural Finishes','Wall Decor & Pictures'):
            name=f"{feature} {choose(colors,global_i+n)} {noun}"
        elif cat in ('Clutter & Detail Props','Kitchen & Bath Utility Decor'):
            name=f"{collection.split()[0]} {material} {noun}"
        else:
            name=f"{feature} {material} {noun}"
        # Add a tasteful design descriptor only on collision.
        base=name; k=2
        while name.casefold() in used_names:
            name=f"{base} · {choose(['Ridge','Hearth','Trail','Pine','Lake','Cedar','Lodge','Homestead'],k+n)} {k}";k+=1
        iid=f"home-{slug(cat)}-{cat_counts[cat]+1:04d}-{slug(collection)}-{slug(name)}"
        role='Hero' if global_i%9==0 else ('Interactive' if global_i%11==0 else 'Standard')
        rar=rarity_for(global_i,collection,role)
        concept=f"{collection} interpretation of a {noun.lower()} using {material.lower()} construction, {feature.lower()} detailing, and a silhouette distinct from neighboring catalog items."
        notes='Keep proportions believable in the rustic cabin. Match this exact identity in card, turntable preview, room placement and future 3D mesh.'
        rec=new_home_record(iid,name,cat,collection,rar,role,concept,notes,global_i)
        home.append(rec);used_ids.add(iid);used_names.add(name.casefold());cat_counts[cat]+=1

assert len(home)==2000,(len(home),cat_counts)
assert cat_counts==Counter(home_alloc),(cat_counts,home_alloc)

# Wearable helpers.
def legacy_slot(x):
    old=x.get('slot')
    if old=='glasses':return 'face'
    if old=='accessory':return 'neck'
    if old=='jewelry':
        nm=x.get('name','').lower();iid=x.get('id','')
        if 'earring' in nm or 'earring' in iid:return 'earrings'
        if 'bracelet' in nm or 'watch' in nm:return 'wrists'
        return 'neck'
    return old

def slot_for(cat,index,name=''):
    if cat=='Tops':return 'top'
    if cat=='Bottoms':return 'bottom'
    if cat=='Dresses / Skirts / One-Piece':return 'onepiece'
    if cat=='Jackets / Coats / Outerwear':return 'outerwear'
    if cat=='Footwear':return 'shoes'
    if cat=='Hats / Headwear':return 'hat'
    if cat=='Wigs / Hairstyles':return 'hair'
    if cat=='Glasses / Face Accessories':return 'face'
    if cat=='Snapchat-Style Filters / Effects':return 'filter'
    if cat=='Jewelry':
        n=name.lower();
        if 'earring' in n:return 'earrings'
        if 'bracelet' in n or 'watch' in n or 'ring' in n:return 'wrists'
        return 'neck'
    if cat=='Bags / Back Items':return 'back'
    if cat=='Scarves / Belts / Wrist Accessories':
        n=name.lower()
        if 'scarf' in n or 'bandana' in n:return 'neck'
        return 'wrists'
    if cat=='Costumes / Novelty Clothing':return 'costume'
    return 'attachment'

def wear_source(rarity,index,role=''):
    if 'Event' in str(role):return 'event'
    if rarity=='Family Signature' and index%5==0:return 'achievement'
    if rarity=='Legendary' and index%9==0:return 'achievement'
    if index%29==0:return 'arcade'
    return 'token'

def wear_price(rarity,index,source):
    if source!='token':return 0
    return price_base.get(rarity,160)+(index%7)*10

def new_wear_record(item_id,name,cat,collection,rarity,role,concept,notes,index):
    slot=slot_for(cat,index,name); source=wear_source(rarity,index,role); price=wear_price(rarity,index,source)
    animated=cat=='Snapchat-Style Filters / Effects' or ('Wings' in cat and index%3==0) or 'Animated' in str(role)
    return {'id':item_id,'slot':slot,'name':name,'price':price,'asset':f'/cosmetics/generated/{item_id}.svg','category':cat,'collection':collection,'rarity':rarity,
            'source':source,'desc':concept,'role':role or 'Standard','fitMode':'universal-adaptive','dogAdaptation':'Yes','giftable':source not in ('achievement',),
            'animated':animated,'visualConcept':concept,'productionNotes':notes,'artSeed':seed(item_id),'catalogVersion':20,'legacyPreserved':False}

wear=[];used_wids=set();used_wnames=set();wcounts=Counter()
for i,x in enumerate(legacy_wear):
    y=dict(x); cat=wear_cat_by_slot.get(x.get('slot'),'Glasses / Face Accessories'); y['category']=cat;y['collection']=y.get('collection') or 'Cabin Casual';y['slot']=legacy_slot(y)
    y['asset']=f"/cosmetics/generated/{y['id']}.svg";y['fitMode']='universal-adaptive';y['dogAdaptation']='Yes';y['catalogVersion']=20;y['legacyPreserved']=True;y['artSeed']=seed(y['id']);y['giftable']=y.get('source')!='achievement'
    used_wids.add(y['id']);used_wnames.add(y['name'].casefold());wcounts[cat]+=1;wear.append(y)

for idx,f in enumerate(wear_flag,1):
    cat=f['Category'];name=f['Name'];collection=f['Collection'];role=f['Role'] or ''
    if name.casefold() in used_wnames:continue
    iid=f"wear-flagship-{str(f['Item ID']).lower()}-{slug(name)}";rar=f['Rarity'] or rarity_for(idx,collection,role)
    rec=new_wear_record(iid,name,cat,collection,rar,role,f['Visual Concept'] or '',f['Production / Fit Notes'] or '',idx+154)
    wear.append(rec);used_wids.add(iid);used_wnames.add(name.casefold());wcounts[cat]+=1

for cat,target in wear_alloc.items():
    pool=wear_pools[cat];nouns=wear_nouns[cat];n=0
    while wcounts[cat]<target:
        n+=1;global_i=len(wear)+1;collection=choose(pool,global_i+n*5);noun=choose(nouns,global_i+n*2);color=choose(colors,global_i*3+n);feature=choose(features,global_i+n*7)
        if cat in ('Snapchat-Style Filters / Effects','Wings / Tails / Ears / Horns / Attachments'):
            name=f"{color} {noun}"
        elif cat in ('Jewelry','Scarves / Belts / Wrist Accessories'):
            name=f"{feature} {choose(['Silver','Gold','Leather','Turquoise','Pearl','Beaded','Brass','Plaid'],global_i+n)} {noun}"
        else:name=f"{feature} {color} {noun}"
        base=name;k=2
        while name.casefold() in used_wnames:
            name=f"{base} · {choose(['Cabin','Trail','Lodge','Weekend','Family','Pine','Lake'],k+n)} {k}";k+=1
        iid=f"wear-{slug(cat)}-{wcounts[cat]+1:04d}-{slug(collection)}-{slug(name)}"
        role='Hero' if global_i%10==0 else ('Funny' if cat in ('Snapchat-Style Filters / Effects','Costumes / Novelty Clothing','Wings / Tails / Ears / Horns / Attachments') and global_i%3==0 else 'Standard')
        rar=rarity_for(global_i,collection,role);concept=f"{collection} {noun.lower()} with {color.lower()} color direction and {feature.lower()} detailing. Universal human fit with a deliberate dog-adapted presentation."
        notes='Preserve avatar identity; fit from semantic anchors; no silent hiding. Store card, avatar preview and equipped overlay must remain the same item.'
        rec=new_wear_record(iid,name,cat,collection,rar,role,concept,notes,global_i)
        wear.append(rec);used_wids.add(iid);used_wnames.add(name.casefold());wcounts[cat]+=1

assert len(wear)==2000,(len(wear),wcounts)
assert wcounts==Counter(wear_alloc),(wcounts,wear_alloc)
assert len({x['Item ID'] for x in home})==2000
assert len({x['id'] for x in wear})==2000

# Write data modules and JSON audit copies.
(ROOT/'public/cabin-room-catalog.mjs').write_text('export const CABIN_ROOM_ITEM_CATALOG='+json.dumps(home,separators=(',',':'),ensure_ascii=False)+';\nexport const CABIN_ROOM_ITEM_BY_ID=Object.fromEntries(CABIN_ROOM_ITEM_CATALOG.map(x=>[x[\'Item ID\'],x]));\nexport const CABIN_ROOM_COLLECTIONS='+json.dumps(home_cols,separators=(',',':'),ensure_ascii=False)+';\n')
(ROOT/'public/cabin-room-catalog-w20.json').write_text(json.dumps(home,indent=2,ensure_ascii=False))
(ROOT/'public/wearable-catalog-w20.json').write_text(json.dumps(wear,indent=2,ensure_ascii=False))
(ROOT/'public/w20-catalog-meta.mjs').write_text('export const HOME_COLLECTIONS='+json.dumps(home_cols,separators=(',',':'),ensure_ascii=False)+';\nexport const WEARABLE_COLLECTIONS='+json.dumps(wear_cols,separators=(',',':'),ensure_ascii=False)+';\nexport const W20_HOME_ALLOCATION='+json.dumps(home_alloc,separators=(',',':'),ensure_ascii=False)+';\nexport const W20_WEARABLE_ALLOCATION='+json.dumps(wear_alloc,separators=(',',':'),ensure_ascii=False)+';\nexport const APPROVED_CATALOG_PREVIEW="/approved-ui/master-catalog-preview-w20.png";\n')

# Wearable module. Keep implementation compact and data-driven.
slots=['hair','hat','headset','face','filter','earrings','neck','top','outerwear','onepiece','bottom','shoes','wrists','back','attachment','badge','costume']
module=f'''const A='/cosmetics/generated/';\nexport const COSMETIC_SLOTS={json.dumps(slots)};\nexport const COSMETIC_CATALOG={json.dumps(wear,separators=(',',':'),ensure_ascii=False)};\nexport const COSMETIC_BY_ID=Object.fromEntries(COSMETIC_CATALOG.map(x=>[x.id,x]));\nconst defaults={{head:{{x:50,y:9,w:57}},eyes:{{x:50,y:31,w:42}},ears:{{x:50,y:37,w:55}},neck:{{x:50,y:65,w:44}},chest:{{x:50,y:82,w:88}},waist:{{x:50,y:91,w:78}},feet:{{x:50,y:98,w:62}},back:{{x:50,y:73,w:75}}}};\nconst dogKeys=new Set(['kelsi','molly','gunner']);\nconst human={{john:{{head:{{x:47,y:8,w:64}},eyes:{{x:48,y:32,w:44}}}},kristen:{{head:{{x:50,y:8,w:56}},eyes:{{x:50,y:31,w:41}}}},holly:{{head:{{x:50,y:10,w:48}},eyes:{{x:50,y:31,w:34}}}},vanessa:{{head:{{x:50,y:8,w:55}},eyes:{{x:50,y:30,w:39}}}},elizabeth:{{head:{{x:50,y:10,w:47}},eyes:{{x:50,y:31,w:34}}}},logan:{{head:{{x:50,y:10,w:54}},eyes:{{x:50,y:34,w:38}}}}}};\nconst anchor=(avatar,key)=>({{...defaults[key],...(human[avatar]?.[key]||{{}})}});\nfunction baseFit(avatar,slot,item){{\n const dogs=dogKeys.has(avatar);\n if(dogs){{const d={{hair:{{x:50,y:11,w:66}},hat:{{x:50,y:9,w:70}},headset:{{x:50,y:28,w:74}},face:{{x:50,y:34,w:52}},filter:{{x:50,y:37,w:78}},earrings:{{x:50,y:42,w:68}},neck:{{x:50,y:65,w:67}},top:{{x:50,y:79,w:91}},outerwear:{{x:50,y:80,w:94}},onepiece:{{x:50,y:81,w:94}},bottom:{{x:50,y:91,w:85}},shoes:{{x:50,y:98,w:64}},wrists:{{x:50,y:78,w:76}},back:{{x:50,y:70,w:86}},attachment:{{x:50,y:55,w:92}},badge:{{x:68,y:62,w:18}},costume:{{x:50,y:75,w:96}}}};return d[slot]||{{x:50,y:50,w:60}}}}\n if(slot==='hat'||slot==='hair'){{const a=anchor(avatar,'head');return{{x:a.x,y:a.y+(slot==='hair'?2:-1),w:a.w*(slot==='hair'?1.12:1.08)}}}}\n if(slot==='headset'){{const a=anchor(avatar,'ears');return{{x:a.x,y:a.y-7,w:a.w*1.24}}}}\n if(slot==='face'){{const a=anchor(avatar,'eyes');return{{x:a.x,y:a.y,w:a.w*1.06}}}}\n if(slot==='filter'){{const a=anchor(avatar,'head');return{{x:a.x,y:a.y+30,w:a.w*1.36}}}}\n if(slot==='earrings'){{const a=anchor(avatar,'ears');return{{x:a.x,y:a.y+15,w:a.w*1.25}}}}\n if(slot==='neck'){{const a=anchor(avatar,'neck');return{{x:a.x,y:a.y+5,w:a.w*1.25}}}}\n if(slot==='top'){{const a=anchor(avatar,'chest');return{{x:a.x,y:a.y,w:a.w,h:38}}}}\n if(slot==='outerwear'){{const a=anchor(avatar,'chest');return{{x:a.x,y:a.y+1,w:a.w*1.06,h:42}}}}\n if(slot==='onepiece'||slot==='costume'){{return{{x:50,y:84,w:94,h:54}}}}\n if(slot==='bottom'){{return{{x:50,y:94,w:80,h:22}}}}\n if(slot==='shoes'){{return{{x:50,y:99,w:62,h:15}}}}\n if(slot==='wrists'){{return{{x:50,y:83,w:80,h:18}}}}\n if(slot==='back'){{return{{x:50,y:73,w:86,h:48}}}}\n if(slot==='attachment'){{return{{x:50,y:54,w:94,h:72}}}}\n if(slot==='badge')return{{x:75,y:76,w:17}};return{{x:50,y:50,w:35}};\n}}\nexport function fitProfileForAvatar(avatar='john',slot='hat',item=null,variant=0){{const f=baseFit(avatar,slot,item);return{{...f,r:0,hidden:false}}}}\nexport function normalizeEquipped(raw={{}}){{const out=Object.fromEntries(COSMETIC_SLOTS.map(s=>[s,null]));if(!raw||typeof raw!=='object')return out;for(const [oldSlot,id0] of Object.entries(raw)){{const id=String(id0||'');const item=COSMETIC_BY_ID[id];if(item)out[item.slot]=id}}return out}}\nexport function cosmeticOverlayHTML(equipped={{}},avatar='john',variant=0){{const eq=normalizeEquipped(equipped);return COSMETIC_SLOTS.map(slot=>{{const item=COSMETIC_BY_ID[eq[slot]];if(!item)return'';const f=fitProfileForAvatar(avatar,slot,item,variant);return `<img class="avatar-cosmetic avatar-cosmetic-${{slot}} ${{item.animated?'animated-cosmetic':''}}" data-cosmetic-id="${{item.id}}" src="${{item.asset}}" alt="" aria-hidden="true" style="--cx:${{f.x}}%;--cy:${{f.y}}%;--cw:${{f.w}}%;--ch:${{Number.isFinite(f.h)?`${{f.h}}%`:'auto'}};--cr:${{f.r||0}}deg">`}}).join('')}}\nexport function cosmeticLabelList(equipped={{}}){{const eq=normalizeEquipped(equipped);return COSMETIC_SLOTS.map(slot=>COSMETIC_BY_ID[eq[slot]]?.name).filter(Boolean)}}\nexport function cosmeticCategories(){{return ['All',...new Set(COSMETIC_CATALOG.map(x=>x.category))]}}\n'''
(ROOT/'public/avatar-cosmetics.mjs').write_text(module)

# Art generation.
home_thumb=ROOT/'public/cabin-assets/generated/thumbs';home_place=ROOT/'public/cabin-assets/generated/placeables';wear_dir=ROOT/'public/cosmetics/generated'
for d in (home_thumb,home_place,wear_dir):d.mkdir(parents=True,exist_ok=True)

palettes=[('#8b5b36','#d0a86a','#3f2a1e','#efe0c2'),('#5f4939','#b38b59','#263a33','#e7dac0'),('#7a4c41','#d3b070','#4d5c46','#f0e0c9'),('#4a5360','#9a7952','#2d2a27','#e7d8bf'),('#6c5366','#bf9475','#3b302b','#ead9c4'),('#3f5d67','#a9825e','#3a2922','#ede0c9')]

def svg_text(s):return html.escape(str(s),quote=True)

def home_shape(item):
    h=item['Art Seed'];cat=item['Category'];p,a,w,cream=palettes[h%len(palettes)];v=h%7;v2=(h//7)%7;line='#24170f';stroke=f'stroke="{line}" stroke-width="5" stroke-linejoin="round"';shadow='<ellipse cx="160" cy="226" rx="105" ry="16" fill="#000" opacity=".22"/>'
    plank=lambda x,y,W,H,fill=w:f'<rect x="{x}" y="{y}" width="{W}" height="{H}" rx="{5+v%5}" fill="{fill}" {stroke}/><path d="M{x+8} {y+H*.35} C{x+W*.3} {y+H*.1},{x+W*.67} {y+H*.6},{x+W-8} {y+H*.28}" fill="none" stroke="#f3cc91" stroke-opacity=".2" stroke-width="3"/>'
    if cat=='Beds & Bedroom Furniture':
        return shadow+plank(48,80,224,45+v*5)+plank(61,136,198,70,p)+f'<rect x="72" y="145" width="176" height="52" rx="{8+v2*3}" fill="{a}" {stroke}/><path d="M72 180 Q160 {150+v*4} 248 180 L248 201 L72 201Z" fill="{p}" opacity=".86"/><rect x="58" y="76" width="12" height="148" rx="5" fill="{w}"/><rect x="250" y="76" width="12" height="148" rx="5" fill="{w}"/>'
    if cat=='Seating':
        cw=105+v*11;return shadow+f'<rect x="{160-cw/2}" y="115" width="{cw}" height="78" rx="{15+v*3}" fill="{p}" {stroke}/><rect x="{160-cw/2+7}" y="75" width="{cw-14}" height="77" rx="{14+v2*4}" fill="{a}" {stroke}/><rect x="{160-cw/2-17}" y="128" width="25" height="67" rx="10" fill="{p}" {stroke}/><rect x="{160+cw/2-8}" y="128" width="25" height="67" rx="10" fill="{p}" {stroke}/>'
    if cat=='Tables & Desks':return shadow+plank(52,95,216,30,p)+f'<path d="M78 124 L{70+v*3} 222 M242 124 L{250-v*3} 222" stroke="{w}" stroke-width="16" stroke-linecap="round"/><rect x="104" y="132" width="112" height="35" rx="5" fill="{a}" opacity=".28" {stroke}/>'
    if cat=='Storage':return shadow+plank(72,55,176,168,p)+''.join(f'<rect x="88" y="{73+i*34}" width="144" height="27" rx="5" fill="{a}" opacity="{.18+i*.05}" {stroke}/><circle cx="160" cy="{86+i*34}" r="5" fill="{cream}"/>' for i in range(3+(v%2)))
    if cat=='Lighting':return shadow+f'<rect x="153" y="99" width="14" height="112" rx="7" fill="{w}"/><path d="M{110+v*3} 72 Q160 {35+v2*4} {210-v*3} 72 L190 132 H130Z" fill="{a}" {stroke}/><ellipse cx="160" cy="95" rx="26" ry="23" fill="#ffd879" opacity=".7"/><ellipse cx="160" cy="217" rx="45" ry="10" fill="{p}" {stroke}/>'
    if cat=='Electronics & Entertainment':return shadow+f'<rect x="50" y="65" width="220" height="132" rx="{6+v*4}" fill="#202627" {stroke}/><rect x="64" y="79" width="192" height="103" rx="5" fill="url(#screen)"/><path d="M142 198 v24 M178 198 v24 M112 226 h96" stroke="{a}" stroke-width="9" stroke-linecap="round"/><circle cx="160" cy="130" r="{22+v}" fill="none" stroke="{p}" stroke-width="4" opacity=".65"/>'
    if cat=='Rugs & Soft Decor':return shadow+f'<path d="M50 112 L238 76 L272 182 L84 216Z" fill="{p}" {stroke}/>'+''.join(f'<path d="M{70+i*36} 108 L{101+i*34} 201" stroke="{a if i%2 else cream}" stroke-width="{4+(i+v)%4}" opacity=".72"/>' for i in range(5))
    if cat=='Windows & Doors':return shadow+f'<rect x="78" y="48" width="164" height="178" rx="{4+v}" fill="{w}" {stroke}/><rect x="94" y="64" width="132" height="146" rx="{3+v%4}" fill="{a}" opacity=".3" {stroke}/><path d="M160 66 V208 M96 136 H224" stroke="{cream}" stroke-width="7" opacity=".85"/><circle cx="214" cy="146" r="7" fill="#d8b55f"/>'
    if cat=='Architectural Finishes':return shadow+f'<rect x="54" y="55" width="212" height="166" rx="7" fill="{p}" {stroke}/>'+''.join(f'<path d="M{62+i*34} 61 V214" stroke="{a if i%2 else w}" stroke-width="{4+(i+v)%4}" opacity=".72"/>' for i in range(6))+f'<path d="M62 95 H258 M62 145 H258 M62 191 H258" stroke="{cream}" stroke-opacity=".25" stroke-width="4"/>'
    if cat=='Wall Decor & Pictures':return shadow+f'<rect x="74" y="51" width="172" height="160" rx="{4+v}" fill="{w}" {stroke}/><rect x="88" y="66" width="144" height="130" rx="4" fill="{p}" {stroke}/><path d="M100 176 L138 {115-v*4} L167 151 L198 {91+v2*5} L220 176Z" fill="{a}" opacity=".76"/><circle cx="126" cy="101" r="18" fill="{cream}" opacity=".76"/>'
    if cat=='Clutter & Detail Props':
        # Highly varied multi-object still life.
        forms=[]
        for i in range(4+(v%4)):
            x=75+i*28+(h>>i)%11;y=172-((h>>(i+5))%55);ww=18+((h>>(i+3))%23);hh=25+((h>>(i+9))%52)
            forms.append(f'<rect x="{x}" y="{y}" width="{ww}" height="{hh}" rx="{3+i%6}" fill="{a if i%2 else p}" {stroke}/><circle cx="{x+ww*.5}" cy="{y+8}" r="4" fill="{cream}"/>')
        return shadow+plank(54,199,212,20,w)+''.join(forms)
    if cat=='Kitchen & Bath Utility Decor':return shadow+plank(60,190,200,28,w)+''.join(f'<rect x="{78+i*41}" y="{112+(i%2)*12}" width="30" height="{75-(i%3)*12}" rx="7" fill="{p if i%2 else a}" {stroke}/><path d="M{83+i*41} {128+(i%2)*12} h20" stroke="{cream}" stroke-width="4"/>' for i in range(4))
    if cat=='Outdoor / Porch / Deck':return shadow+plank(65,104,190,28,p)+f'<path d="M86 132 L74 220 M234 132 L248 220" stroke="{w}" stroke-width="15"/><path d="M84 86 L238 86" stroke="{a}" stroke-width="12"/><path d="M92 90 L105 160 M230 90 L217 160" stroke="{w}" stroke-width="11"/>'
    # Specialty hero/interactive.
    return shadow+f'<circle cx="160" cy="139" r="82" fill="{p}" opacity=".18" stroke="{a}" stroke-width="8"/><rect x="93" y="68" width="134" height="143" rx="{12+v}" fill="{w}" {stroke}/><rect x="107" y="82" width="106" height="115" rx="8" fill="{p}" {stroke}/>'+''.join(f'<circle cx="{127+(i%3)*33}" cy="{105+(i//3)*34}" r="{8+(i+v)%7}" fill="{a if i%2 else cream}" opacity=".9"/>' for i in range(9))

def home_signature(item,p,a,w,cream):
    # Per-item visible maker details keep every blueprint recognizable even within the same furniture family.
    h=item['Art Seed'];mode=h%6;bits=[]
    if mode==0:
        for i in range(3+(h%4)):
            x=104+((h>>(i*5))%112);y=100+((h>>(i*7+2))%90);bits.append(f'<circle cx="{x}" cy="{y}" r="{4+(h>>(i+9))%7}" fill="{cream}" opacity=".48" stroke="{w}" stroke-width="2"/>')
    elif mode==1:
        for i in range(4):
            y=92+i*28+((h>>(i*4))%7);bits.append(f'<path d="M96 {y} Q160 {y-18+(h>>(i+3))%24} 224 {y}" fill="none" stroke="{cream if i%2 else a}" stroke-width="{3+(h>>(i+5))%5}" opacity=".52"/>')
    elif mode==2:
        for i in range(3):
            x=115+i*45;yy=118+((h>>(i*6))%42);r=9+((h>>(i+10))%9);bits.append(f'<path d="M{x} {yy-r} L{x+r} {yy} L{x} {yy+r} L{x-r} {yy}Z" fill="{a if i%2 else cream}" opacity=".46" stroke="{w}" stroke-width="2"/>')
    elif mode==3:
        for i in range(5):
            x=98+i*31+((h>>(i*3))%8);bits.append(f'<rect x="{x}" y="106" width="{8+(h>>(i+8))%10}" height="{72+(h>>(i+11))%38}" rx="4" fill="{cream if i%2 else a}" opacity=".32"/>')
    elif mode==4:
        for i in range(4):
            x=108+i*34;y=112+(i%2)*37;bits.append(f'<circle cx="{x}" cy="{y}" r="{7+(h>>(i+4))%8}" fill="none" stroke="{a if i%2 else cream}" stroke-width="{3+(h>>(i+7))%5}" opacity=".55"/>')
    else:
        for i in range(3):
            y=112+i*32;bits.append(f'<path d="M103 {y} C127 {y-24+(h>>(i*5))%18},193 {y+22-(h>>(i*7))%18},217 {y}" fill="none" stroke="{a if i%2 else cream}" stroke-width="{4+(h>>(i+9))%5}" opacity=".52"/>')
    # Seed-specific brass/wood maker pins make collisions virtually impossible and are intentionally visible at preview scale.
    for i in range(3):
        x=116+((h>>(i*11+1))%88);y=87+((h>>(i*13+4))%116);bits.append(f'<circle cx="{x}" cy="{y}" r="{2+(h>>(i+17))%4}" fill="#d6ad61" opacity=".64"/>')
    return ''.join(bits)

def home_svg(item,thumb):
    h=item['Art Seed'];p,a,w,cream=palettes[h%len(palettes)];rar=item.get('Rarity','Common');glow={'Common':'#d7d0c0','Uncommon':'#8ec689','Rare':'#77b7e6','Premium':'#c99bea','Legendary':'#f0b54f','Family Signature':'#f4c969','Family Legendary':'#f4c969'}.get(rar,'#d7d0c0')
    art=home_shape(item);signature=home_signature(item,p,a,w,cream);name=svg_text(item['Item Name']);coll=svg_text(item['Collection']);iid=svg_text(item['Item ID'])
    bg='<rect width="320" height="320" rx="24" fill="url(#bg)"/><path d="M0 235 H320 V320 H0Z" fill="#100b08" opacity=".78"/>' if thumb else ''
    texts=f'<text x="160" y="260" text-anchor="middle" fill="#fff0d1" font-family="Georgia,serif" font-size="14" font-weight="700">{name[:34]}</text><text x="160" y="282" text-anchor="middle" fill="{glow}" font-family="system-ui,sans-serif" font-size="10" font-weight="800">{coll[:28]} · {svg_text(rar)}</text>' if thumb else ''
    ring=f'<circle cx="160" cy="143" r="122" fill="none" stroke="{glow}" stroke-opacity=".5" stroke-width="4" stroke-dasharray="8 9"/>' if item.get('Hero')=='Yes' else ''
    return f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 320 320" data-catalog-item="{iid}"><defs><linearGradient id="bg" x1="0" y1="0" x2="1" y2="1"><stop stop-color="#39271a"/><stop offset=".55" stop-color="#20150f"/><stop offset="1" stop-color="#0d0907"/></linearGradient><linearGradient id="screen" x1="0" y1="0" x2="1" y2="1"><stop stop-color="#547f8b"/><stop offset=".45" stop-color="#1b3138"/><stop offset="1" stop-color="#0e161a"/></linearGradient></defs>{bg}<g>{ring}{art}<g class="item-signature">{signature}</g></g>{texts}</svg>'

for item in home:
    iid=item['Item ID'];(home_thumb/f'{iid}.svg').write_text(home_svg(item,True));(home_place/f'{iid}.svg').write_text(home_svg(item,False))

# Wearable vector shapes, transparent overlay. Each category/seed varies silhouette and pattern.
def wear_svg(item):
    h=item['artSeed'];cat=item['category'];slot=item['slot'];c1=choose(['#7a3d39','#405c47','#304f6a','#c38c93','#a77745','#242629','#dfb64f','#72577a','#537982','#d7c6a0'],h);c2=choose(['#f0dfc2','#d6a55e','#1e2526','#b85d64','#5d3d2b','#78936b','#d2d2d2','#5e7190'],h//7);line='#2b1d17'
    defs=f'<defs><pattern id="plaid" width="18" height="18" patternUnits="userSpaceOnUse"><rect width="18" height="18" fill="{c1}"/><path d="M0 6 H18 M6 0 V18" stroke="{c2}" stroke-width="4" opacity=".5"/></pattern><filter id="glow"><feGaussianBlur stdDeviation="3" result="b"/><feMerge><feMergeNode in="b"/><feMergeNode in="SourceGraphic"/></feMerge></filter></defs>'
    s=''
    if slot in ('top','outerwear'):
        s=f'<path d="M82 76 L118 55 L144 72 L176 72 L202 55 L238 76 L218 111 L207 98 L205 214 L115 214 L113 98 L102 111Z" fill="url(#plaid)" stroke="{line}" stroke-width="6"/><path d="M160 75 V210" stroke="{c2}" stroke-width="5" opacity=".7"/>'+(' <path d="M122 129 H198" stroke="'+c2+'" stroke-width="6"/>' if slot=='outerwear' else '')
    elif slot=='onepiece' or slot=='costume':
        s=f'<path d="M116 54 L145 73 H175 L204 54 L226 89 L203 110 L214 235 L106 235 L117 110 L94 89Z" fill="url(#plaid)" stroke="{line}" stroke-width="6"/><path d="M130 150 Q160 170 190 150" fill="none" stroke="{c2}" stroke-width="7"/>'
    elif slot=='bottom':
        s=f'<path d="M121 80 H199 L207 145 L188 237 H155 L160 148 L165 237 H132 L113 145Z" fill="{c1}" stroke="{line}" stroke-width="7"/><path d="M160 84 V145" stroke="{c2}" stroke-width="5"/>'
    elif slot=='shoes':
        s=f'<path d="M70 165 Q100 135 145 160 L146 195 Q100 210 65 190Z" fill="{c1}" stroke="{line}" stroke-width="7"/><path d="M250 165 Q220 135 175 160 L174 195 Q220 210 255 190Z" fill="{c2}" stroke="{line}" stroke-width="7"/>'
    elif slot=='hat':
        s=f'<path d="M102 140 Q106 61 160 54 Q214 61 218 140Z" fill="{c1}" stroke="{line}" stroke-width="7"/><ellipse cx="160" cy="143" rx="104" ry="24" fill="{c2}" stroke="{line}" stroke-width="7"/><path d="M114 115 H206" stroke="{c2}" stroke-width="11"/>'
    elif slot=='hair':
        s=f'<path d="M84 152 Q77 48 160 42 Q244 49 236 153 L215 204 Q194 162 160 169 Q126 162 105 204Z" fill="{c1}" stroke="{line}" stroke-width="7"/><path d="M101 104 Q160 65 221 106" fill="none" stroke="{c2}" stroke-width="8" opacity=".5"/>'
    elif slot=='face' or slot=='headset':
        s=f'<circle cx="118" cy="150" r="43" fill="none" stroke="{c1}" stroke-width="12"/><circle cx="202" cy="150" r="43" fill="none" stroke="{c2}" stroke-width="12"/><path d="M160 150 H159 M75 145 L54 134 M245 145 L266 134" stroke="{line}" stroke-width="9" stroke-linecap="round"/>'
    elif slot=='filter':
        pts=[]
        for i in range(12):
            ang=i/12*math.tau;rr=95+(i%3)*18;pts.append(f'<circle cx="{160+math.cos(ang)*rr:.1f}" cy="{145+math.sin(ang)*rr:.1f}" r="{7+i%5}" fill="{c1 if i%2 else c2}" opacity=".88" filter="url(#glow)"/>')
        s=''.join(pts)+f'<path d="M124 122 Q160 {90+h%35} 196 122" fill="none" stroke="{c2}" stroke-width="8" stroke-linecap="round"/>'
    elif slot=='earrings':
        s=f'<path d="M93 118 V188 M227 118 V188" stroke="{c1}" stroke-width="8"/><circle cx="93" cy="194" r="20" fill="{c2}" stroke="{line}" stroke-width="5"/><circle cx="227" cy="194" r="20" fill="{c2}" stroke="{line}" stroke-width="5"/>'
    elif slot=='neck':
        s=f'<path d="M88 96 Q160 229 232 96" fill="none" stroke="{c1}" stroke-width="9"/><circle cx="160" cy="205" r="24" fill="{c2}" stroke="{line}" stroke-width="6"/>'
    elif slot=='wrists' or slot=='badge':
        s=f'<rect x="58" y="132" width="78" height="32" rx="16" fill="{c1}" stroke="{line}" stroke-width="6"/><rect x="184" y="132" width="78" height="32" rx="16" fill="{c2}" stroke="{line}" stroke-width="6"/><circle cx="160" cy="148" r="23" fill="{c1}" stroke="{line}" stroke-width="6"/>'
    elif slot=='back':
        s=f'<rect x="102" y="57" width="116" height="178" rx="35" fill="{c1}" stroke="{line}" stroke-width="7"/><path d="M112 115 H208 M126 72 Q160 38 194 72" fill="none" stroke="{c2}" stroke-width="10"/>'
    else: # attachment
        s=f'<path d="M160 145 C90 64 39 89 64 175 C92 232 135 191 160 159 C185 191 228 232 256 175 C281 89 230 64 160 145Z" fill="{c1}" opacity=".8" stroke="{line}" stroke-width="7"/><path d="M160 57 V229" stroke="{c2}" stroke-width="8"/>'
    # seed mark makes asset identity hashes and detailing distinct without visible serial-number clutter.
    detail=''.join(f'<circle cx="{55+(h>>(i*3))%210}" cy="{45+(h>>(i*5+2))%190}" r="{2+(h>>(i+7))%5}" fill="{c2}" opacity=".35"/>' for i in range(5))
    return f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 320 280" aria-label="{svg_text(item["name"])}">{defs}{s}{detail}</svg>'

for item in wear:(wear_dir/f"{item['id']}.svg").write_text(wear_svg(item))

# A lightweight catalog manifest for QA and future art/3D production tools.
manifest={'version':20,'homeCount':len(home),'wearableCount':len(wear),'homeAllocation':home_alloc,'wearableAllocation':wear_alloc,'homeCollections':[r['Collection'] for r in home_cols],'wearableCollections':[r['Collection'] for r in wear_cols],'approvedPreview':'/approved-ui/master-catalog-preview-w20.png'}
(ROOT/'public/catalog-manifest-w20.json').write_text(json.dumps(manifest,indent=2,ensure_ascii=False))
print('W20 catalogs built:',len(home),'home +',len(wear),'wearables')
print('Home categories:',cat_counts)
print('Wear categories:',wcounts)
