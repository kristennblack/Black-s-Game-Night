from pathlib import Path
from openpyxl import load_workbook
import json, html, hashlib, shutil, re
ROOT=Path('/mnt/data/bfgn_w21')
SRC=Path('/mnt/data/Black_Family_Game_Night_V1_World_Props_Master_Catalog.xlsx')
shutil.copy2(SRC, ROOT/'Black_Family_Game_Night_V1_World_Props_Master_Catalog.xlsx')
wb=load_workbook(SRC,read_only=True,data_only=True)
ws=wb['All 2000 Props']
headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
rows=[]
for vals in ws.iter_rows(min_row=2,values_only=True):
    rec={headers[i]:vals[i] for i in range(len(headers))}
    rec={k:(v if v is not None else '') for k,v in rec.items()}
    rows.append(rec)
assert len(rows)==2000
assert len({r['Prop ID'] for r in rows})==2000
assert len({r['Prop Name'] for r in rows})==2000
out=ROOT/'public/world-props/generated';out.mkdir(parents=True,exist_ok=True)
# Warm approved catalog palette with per-record signatures.
PALETTES=[('#2a1a12','#b57b43','#e4c184','#6e4931'),('#13241a','#66835b','#d3b676','#342319'),('#211817','#855142','#dbc28e','#4b3527'),('#202024','#657287','#d4bc86','#463528'),('#2b2417','#96723f','#ead3a0','#5e432d')]

def seed_for(r): return int(hashlib.sha256((str(r['Prop ID'])+'|'+str(r['Prop Name'])+'|'+str(r['Collection'])).encode()).hexdigest()[:8],16)
def esc(s): return html.escape(str(s),quote=True)
def wrap(s,n=22):
    words=str(s).split(); lines=[]; cur=''
    for w in words:
        if len(cur)+len(w)+1>n and cur: lines.append(cur); cur=w
        else: cur=(cur+' '+w).strip()
    if cur: lines.append(cur)
    return lines[:2]

def silhouette(r,seed):
    cat=str(r['Category']); sub=str(r['Subcategory']).lower(); accent=PALETTES[seed%len(PALETTES)][1]; wood=PALETTES[seed%len(PALETTES)][3]
    xj=(seed%19)-9; yj=((seed>>5)%13)-6
    # Item-specific, readable silhouette families. Seeded details keep every generated identity unique.
    if 'Workshop' in cat:
        if any(k in sub for k in ['tire','wheel']): return f'<ellipse cx="160" cy="142" rx="66" ry="66" fill="#191817" stroke="{accent}" stroke-width="23"/><ellipse cx="160" cy="142" rx="24" ry="24" fill="#3d342b"/><path d="M104 201h112v15H104z" fill="{wood}"/>'
        if any(k in sub for k in ['tool','chest','box']): return f'<rect x="78" y="92" width="164" height="111" rx="12" fill="{accent}" stroke="#2a211b" stroke-width="8"/><path d="M111 90v-18h98v18M93 130h134M93 158h134" fill="none" stroke="#dbc58d" stroke-width="7"/><circle cx="160" cy="145" r="7" fill="#2c211a"/>'
        return f'<rect x="83" y="102" width="154" height="101" rx="8" fill="{wood}"/><path d="M101 116h118M112 84l18 34m80-34-18 34" stroke="{accent}" stroke-width="11"/><circle cx="112" cy="200" r="17" fill="#222"/><circle cx="208" cy="200" r="17" fill="#222"/>'
    if 'Farm' in cat:
        if any(k in sub for k in ['hay','bale','straw']): return f'<rect x="79" y="94" width="162" height="105" rx="20" fill="#c79a45" stroke="#8f6630" stroke-width="7"/><path d="M92 117l137 59M100 177l125-55M160 96v101" stroke="#efd28a" stroke-width="5" opacity=".75"/>'
        return f'<path d="M88 102h144v93H88z" fill="{accent}"/><path d="M105 102v93m109-93v93M88 139h144" stroke="#e1c694" stroke-width="7"/><ellipse cx="160" cy="205" rx="82" ry="9" fill="#0004"/>'
    if 'Outdoor' in cat:
        if any(k in sub for k in ['chair','seat']): return f'<path d="M104 89l26 115m86-115-26 115M118 145h84M120 145l-27 58m109-58 27 58" fill="none" stroke="{accent}" stroke-width="14"/><path d="M119 94h83v61h-83z" fill="#5b6f4a"/>'
        return f'<rect x="91" y="105" width="138" height="90" rx="18" fill="{accent}"/><rect x="108" y="86" width="104" height="25" rx="10" fill="#dac79d"/><path d="M111 137h98" stroke="#33251c" stroke-width="7"/>'
    if 'Kitchen' in cat:
        if any(k in sub for k in ['mug','cup']): return f'<path d="M103 97h101v103H103z" fill="#d6c39d" stroke="{accent}" stroke-width="8"/><path d="M204 122c57 0 49 60 0 60" fill="none" stroke="{accent}" stroke-width="13"/><path d="M124 128h58" stroke="#6b4b31" stroke-width="6"/>'
        return f'<ellipse cx="160" cy="111" rx="61" ry="22" fill="#cbb892"/><path d="M99 111v73c0 30 122 30 122 0v-73" fill="{accent}"/><ellipse cx="160" cy="184" rx="61" ry="22" fill="#95734b"/>'
    if 'Pet' in cat:
        return f'<path d="M93 180c0-45 32-74 67-74s67 29 67 74c0 28-134 28-134 0z" fill="{accent}"/><circle cx="132" cy="107" r="20" fill="{wood}"/><circle cx="188" cy="107" r="20" fill="{wood}"/><path d="M126 181c17-24 51-24 68 0" fill="none" stroke="#e9d8ad" stroke-width="8"/>'
    if 'Wall' in cat:
        return f'<rect x="89" y="74" width="142" height="132" rx="5" fill="{wood}" stroke="{accent}" stroke-width="13"/><rect x="107" y="92" width="106" height="96" fill="#bda97f"/><path d="M119 171l30-42 19 25 22-31 14 48z" fill="#5d7557"/><circle cx="185" cy="113" r="13" fill="#e5bf69"/>'
    if 'Interactive' in cat:
        return f'<rect x="83" y="82" width="154" height="122" rx="15" fill="{wood}" stroke="{accent}" stroke-width="9"/><rect x="101" y="101" width="118" height="66" rx="8" fill="#16333a"/><circle cx="127" cy="187" r="8" fill="#f0c261"/><circle cx="160" cy="187" r="8" fill="#83c588"/><circle cx="193" cy="187" r="8" fill="#d06b58"/>'
    if 'Seasonal' in cat:
        return f'<path d="M160 69l22 43 48 7-35 34 8 48-43-23-43 23 8-48-35-34 48-7z" fill="{accent}" stroke="#f0d08c" stroke-width="6"/><circle cx="160" cy="140" r="28" fill="{wood}"/>'
    if 'Family' in cat:
        return f'<rect x="88" y="78" width="144" height="126" rx="17" fill="{accent}"/><path d="M115 179v-62h90v62M132 117V96h56v21" fill="none" stroke="#f1ddad" stroke-width="11"/><circle cx="160" cy="142" r="19" fill="{wood}"/>'
    # Indoor cabin default: variable stacked household object.
    return f'<rect x="82" y="111" width="156" height="85" rx="12" fill="{wood}"/><rect x="101" y="92" width="118" height="29" rx="6" fill="{accent}"/><rect x="112" y="73" width="96" height="25" rx="5" fill="#d5be8b"/><path d="M101 {151+yj}h118M{139+xj} 92v104" stroke="#e5cf9b" stroke-width="5" opacity=".72"/>'

def svg(r):
    seed=seed_for(r); bg,accent,gold,wood=PALETTES[seed%len(PALETTES)]; title=wrap(r['Prop Name']); sig=seed%997
    labels=''.join(f'<text x="160" y="{247+i*18}" text-anchor="middle" fill="#f5e5bd" font-family="Georgia,serif" font-size="{14 if i==0 else 12}" font-weight="700">{esc(line)}</text>' for i,line in enumerate(title))
    badges=[]
    if str(r['Hideable'])=='Yes': badges.append('HIDE')
    if str(r['Interactive'])=='Yes': badges.append('LIVE')
    if str(r['Hero'])=='Yes': badges.append('HERO')
    badge=''.join(f'<text x="{20+i*55}" y="29" fill="#f6d277" font-family="Arial,sans-serif" font-size="10" font-weight="800">{b}</text>' for i,b in enumerate(badges[:4]))
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 320 300" data-world-prop="{esc(r['Prop ID'])}" data-signature="{sig}"><defs><linearGradient id="bg" x1="0" y1="0" x2="1" y2="1"><stop stop-color="{bg}"/><stop offset="1" stop-color="#090705"/></linearGradient><filter id="ds"><feDropShadow dx="0" dy="8" stdDeviation="7" flood-opacity=".55"/></filter></defs><rect width="320" height="300" rx="18" fill="url(#bg)"/><rect x="10" y="10" width="300" height="280" rx="14" fill="none" stroke="{gold}" stroke-opacity=".55"/>{badge}<g filter="url(#ds)">{silhouette(r,seed)}</g>{labels}<text x="160" y="286" text-anchor="middle" fill="#bfa77f" font-family="Arial,sans-serif" font-size="8">{esc(r['Collection'])} · {esc(r['Rarity'])} · {esc(r['Prop ID'])}</text></svg>'''

manifest=[]
for r in rows:
    fn=f"{r['Prop ID']}.svg"
    (out/fn).write_text(svg(r),encoding='utf-8')
    rr=dict(r);rr['Art Path']=f"/world-props/generated/{fn}";rr['Art Seed']=seed_for(r);manifest.append(rr)
(ROOT/'public/world-prop-catalog-w21.json').write_text(json.dumps(manifest,ensure_ascii=False,separators=(',',':')),encoding='utf-8')
# Module is static so all games can reuse one identity source.
(ROOT/'public/world-prop-catalog.mjs').write_text('export const WORLD_PROP_CATALOG='+json.dumps(manifest,ensure_ascii=False,separators=(',',':'))+';\nexport const WORLD_PROP_BY_ID=Object.fromEntries(WORLD_PROP_CATALOG.map(x=>[x["Prop ID"],x]));\nexport const WORLD_PROP_COUNT=WORLD_PROP_CATALOG.length;\nexport const WORLD_PROP_FLAGSHIPS=WORLD_PROP_CATALOG.filter(x=>x.Flagship==="Yes");\n',encoding='utf-8')
meta={'version':21,'count':2000,'flagshipCount':sum(r['Flagship']=='Yes' for r in rows),'categories':{},'collections':sorted({str(r['Collection']) for r in rows}),'primaryMaps':sorted({str(r['Primary Map']) for r in rows})}
for r in rows: meta['categories'][r['Category']]=meta['categories'].get(r['Category'],0)+1
(ROOT/'public/world-prop-manifest-w21.json').write_text(json.dumps(meta,indent=2,ensure_ascii=False),encoding='utf-8')
print(json.dumps(meta,indent=2))
